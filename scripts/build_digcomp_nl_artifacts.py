#!/usr/bin/env python3
# versie 5 - 20260417
"""
build_digcomp_nl_artifacts.py

Bouw Nederlandstalige leveringen vanuit vertaalde CSV's in ./digcomp3-l10n/locale:

  a) XLSX – twee varianten:
       *_tweetalig.xlsx  : NL-kolommen vóór EN-kolommen, Nederlandse bladnamen
       *_nl.xlsx         : alleen NL-kolommen, Nederlandse bladnamen
  b) JSON-LD – Nederlandse veldnamen, Verklarende woordenlijst vertaald, bronvermelding

Verbeteringen t.o.v. v4
- Alle DOCX-gerelateerde code verwijderd
- XLSX: NL-kolommen staan VOOR de Engelstalige kolommen
- XLSX: tweede variant met uitsluitend NL-tekst
- XLSX: Leesmij-tabblad overgenomen uit ReadMe_NL.xlsx
- XLSX: consistente Calibri-10-opmaak op alle cellen
- XLSX: bladnamen in het Nederlands
- JSON-LD: veldnamen vernederlandst (name_nl → naam, description_nl → omschrijving enz.)
- JSON-LD: Verklarende woordenlijst vertaald
- JSON-LD: bronvermelding toegevoegd op het hoogste niveau

Gebruik (vanuit de digcomp3-l10n map):
  python scripts/build_digcomp_nl_artifacts.py --build xlsx \
      --src-xlsx "sources/DigComp 3.0 Data Supplement 24 Nov 2025.xlsx" \
      --readme-nl-xlsx "sources/ReadMe_NL.xlsx" --out-dir output
  python scripts/build_digcomp_nl_artifacts.py --build jsonld \
      --src-jsonld "sources/DigComp 3.0 Data Supplement 24 Nov 2025.jsonld" \
      --out-dir output

CSV-formaat verwachte kolommen: location, source, target, context
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter



# ----------------------------
# CSV translations laden
# ----------------------------

def load_component_csv(path: Path) -> Dict[str, Dict[str, str]]:
    """Geeft mapping: location -> rij-dict."""
    out: Dict[str, Dict[str, str]] = {}
    if not path.exists():
        return out
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            loc = (row.get("location") or "").strip()
            if not loc:
                continue
            out[loc] = {
                "source": row.get("source", "") or "",
                "target": row.get("target", "") or "",
                "context": row.get("context", "") or "",
            }
    return out


def load_translations(repo_root: Path) -> Dict[str, Dict[str, Dict[str, str]]]:
    """Laad alle locale-componenten vanuit ./digcomp3-l10n/locale/*/{en.csv,nl.csv}."""
    base = repo_root / "digcomp3-l10n" / "locale"
    tx: Dict[str, Dict[str, Dict[str, str]]] = {}
    if not base.exists():
        raise FileNotFoundError(f"Locale-map niet gevonden: {base}")

    for comp_dir in sorted([p for p in base.iterdir() if p.is_dir()]):
        comp = comp_dir.name
        en = load_component_csv(comp_dir / "en.csv")
        nl = load_component_csv(comp_dir / "nl.csv")
        merged: Dict[str, Dict[str, str]] = {}
        for k in set(en.keys()) | set(nl.keys()):
            merged[k] = {
                "source": (en.get(k) or {}).get("source", "") or (nl.get(k) or {}).get("source", ""),
                "target": (nl.get(k) or {}).get("target", ""),
                "context": (en.get(k) or {}).get("context", "") or (nl.get(k) or {}).get("context", ""),
            }
        tx[comp] = merged
    return tx


def tr(comp: Dict[str, Dict[str, str]], key: str) -> str:
    """Nederlandse vertaling met Engelse fallback."""
    row = comp.get(key)
    if not row:
        return ""
    tgt = (row.get("target") or "").strip()
    if tgt:
        return tgt
    return (row.get("source") or "").strip()


def en(comp: Dict[str, Dict[str, str]], key: str) -> str:
    """Engelstalige brontekst."""
    row = comp.get(key)
    if not row:
        return ""
    return (row.get("source") or "").strip()


# ----------------------------
# Hulpfuncties
# ----------------------------

_num_re = re.compile(r"^\d+(\.\d+)?$")


def norm_num(x) -> str:
    if x is None:
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if x.is_integer():
            return str(int(x))
        return f"{x:.10f}".rstrip("0").rstrip(".")
    s = str(x).strip()
    if re.match(r"^\d+\.0$", s):
        return s[:-2]
    return s


def slugify_term(term: str) -> str:
    s = term.strip().lower()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def build_levels_name_map(levels: Dict[str, Dict[str, str]]) -> Dict[str, str]:
    """Engelstalige vierniveau-naam → Nederlandse vertaling."""
    m: Dict[str, str] = {}
    for k, row in levels.items():
        if k.endswith(".four_level_name"):
            src = (row.get("source") or "").strip()
            tgt = (row.get("target") or "").strip()
            if src and tgt:
                m[src.lower()] = tgt
    return m


# ----------------------------
# Opmaak-hulpfuncties
# ----------------------------

FONT_HEADER = Font(name="Calibri", size=10, bold=True)
FONT_DATA   = Font(name="Calibri", size=10, bold=False)
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")  # lichtblauw voor koprij
ALIGN_WRAP  = Alignment(wrap_text=True, vertical="top")


def style_ws(ws) -> None:
    """Pas consistente Calibri-10 opmaak toe op alle cellen."""
    for row in ws.iter_rows():
        for cell in row:
            is_header = (cell.row == 1)
            cell.font = FONT_HEADER if is_header else FONT_DATA
            cell.alignment = ALIGN_WRAP
            if is_header:
                cell.fill = FILL_HEADER


def write_rows(ws, headers: List[str], rows: List[List]) -> None:
    """Schrijf koprij + gegevensrijen naar ws."""
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_ws(ws)
    # Kolombreedtes automatisch aanpassen (max 60)
    for col_idx, _ in enumerate(headers, start=1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(60, len(str(cell.value))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


# ----------------------------
# Gesloten vocabulaires (vaste vertaaltabellen)
# ----------------------------

# AI-label (hoofdlettergevoelig; strip voor zekerheid)
_AI_LABEL: Dict[str, str] = {
    "ai-implicit":                   "AI-impliciet",
    "ai-explicit":                   "AI-expliciet",
    "ai not implicit or explicit":   "AI niet impliciet of expliciet",
}

# Knowledge / Skill / Attitude
_KSA: Dict[str, str] = {
    "knowledge": "Kennis",
    "skill":     "Vaardigheid",
    "attitude":  "Attitude",
}


def tr_ai(val: str) -> str:
    return _AI_LABEL.get(val.strip().lower(), val.strip())


def tr_ksa(val: str) -> str:
    return _KSA.get(val.strip().lower(), val.strip())


# ----------------------------
# XLSX bouwen
# ----------------------------

def copy_leesmij_sheet(wb: Workbook, readme_nl_xlsx: Path) -> None:
    """Kopieer het LEESMIJ-tabblad uit ReadMe_NL.xlsx naar wb."""
    if not readme_nl_xlsx or not readme_nl_xlsx.exists():
        return
    src_wb = load_workbook(readme_nl_xlsx)
    for sheet_name in ("LEESMIJ", "Leesmij", "leesmij"):
        if sheet_name in src_wb.sheetnames:
            src_ws = src_wb[sheet_name]
            break
    else:
        return  # niet gevonden

    ws = wb.create_sheet("Leesmij", 0)  # als eerste tabblad
    for row in src_ws.iter_rows():
        for cell in row:
            tgt = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.font:
                tgt.font = Font(
                    name=cell.font.name or "Calibri",
                    size=cell.font.size or 10,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                )
            if cell.alignment:
                tgt.alignment = Alignment(
                    wrap_text=cell.alignment.wrap_text,
                    vertical=cell.alignment.vertical,
                )

    # Samengevoegde cellen overnemen
    for merged_range in src_ws.merged_cells.ranges:
        ws.merge_cells(str(merged_range))

    # Kolombreedtes en rijhoogtes overnemen
    for col_letter, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            ws.row_dimensions[row_idx].height = dim.height


def build_xlsx(
    src_xlsx: Path,
    out_dir: Path,
    tx: Dict[str, Dict[str, Dict[str, str]]],
    readme_nl_xlsx: Path | None = None,
    base_name: str = "DigComp_3_0_Data_Supplement",
) -> None:
    """Bouw twee XLSX-varianten: tweetalig (NL+EN) en alleen-NL."""

    wb_src = load_workbook(src_xlsx, data_only=True)

    core   = tx.get("core-framework", {})
    levels = tx.get("levels", {})
    stmts  = tx.get("statements", {})
    outs   = tx.get("outcomes", {})
    gloss  = tx.get("glossary", {})

    level_name_map = build_levels_name_map(levels)

    # --- Gegevens per blad verzamelen ---

    # Blad 1: Competentiegebieden & Competenties
    ws1 = wb_src["1 Competence Areas&Competences"]
    rows1: List[Tuple] = []
    for r in range(2, ws1.max_row + 1):
        area_no  = norm_num(ws1.cell(r, 1).value)
        comp_no  = norm_num(ws1.cell(r, 4).value)
        if not area_no and not comp_no:
            continue
        rows1.append((
            ws1.cell(r, 1).value,                                         # Nummer competentiegebied
            tr(core, f"digcomp.area.{area_no}.label"),                    # NL area naam
            ws1.cell(r, 2).value,                                         # EN area naam
            tr(core, f"digcomp.area.{area_no}.description"),              # NL area omschrijving
            ws1.cell(r, 3).value,                                         # EN area omschrijving
            ws1.cell(r, 4).value,                                         # Nummer competentie
            tr(core, f"digcomp.competence.{comp_no}.label"),              # NL comp naam
            ws1.cell(r, 5).value,                                         # EN comp naam
            tr(core, f"digcomp.competence.{comp_no}.description"),        # NL comp omschrijving
            ws1.cell(r, 6).value,                                         # EN comp omschrijving
        ))

    hdrs1_bi = [
        "Nummer competentiegebied",
        "Naam competentiegebied", "Competence area name",
        "Omschrijving competentiegebied", "Competence area descriptor",
        "Nummer competentie",
        "Naam competentie", "Competence name",
        "Omschrijving competentie", "Competence descriptor",
    ]
    hdrs1_nl = [
        "Nummer competentiegebied",
        "Naam competentiegebied", "Omschrijving competentiegebied",
        "Nummer competentie", "Naam competentie", "Omschrijving competentie",
    ]
    # NL-only kolomindices (0-based) uit de volledige tweetalige rij
    cols1_nl = [0, 1, 3, 5, 6, 8]

    # Blad 2: Beheersingsniveaus
    ws2 = wb_src["2 Proficiency Levels"]
    rows2: List[Tuple] = []
    for r in range(2, ws2.max_row + 1):
        eight_map = ws2.cell(r, 4).value
        if eight_map is None:
            continue
        eight_s = norm_num(eight_map)
        rows2.append((
            tr(levels, f"digcomp.level.{eight_s}.four_level_name"),        # NL niveau naam
            ws2.cell(r, 1).value,                                          # EN niveau naam
            tr(levels, f"digcomp.level.{eight_s}.four_level_description"), # NL vier-niveau omschrijving
            ws2.cell(r, 2).value,                                          # EN vier-niveau omschrijving
            tr(levels, f"digcomp.level.{eight_s}.applies_to"),             # NL doel
            ws2.cell(r, 3).value,                                          # EN doel
            ws2.cell(r, 4).value,                                          # Mapping acht niveaus
            ws2.cell(r, 5).value,                                          # Mapping zes niveaus
            tr(levels, f"digcomp.level.{eight_s}.eight_level_description"),# NL acht-niveau omschrijving
            ws2.cell(r, 6).value,                                          # EN acht-niveau omschrijving
        ))

    hdrs2_bi = [
        "Naam beheersingsniveau", "Proficiency level name",
        "Beschrijving beheersingsniveau (vier niveaus)", "Four level proficiency level description",
        "Doel", "Purpose",
        "Mapping acht niveaus", "Mapping zes niveaus",
        "Beschrijving beheersingsniveau (acht niveaus)", "Eight level proficiency level description",
    ]
    hdrs2_nl = [
        "Naam beheersingsniveau",
        "Beschrijving beheersingsniveau (vier niveaus)",
        "Doel",
        "Mapping acht niveaus", "Mapping zes niveaus",
        "Beschrijving beheersingsniveau (acht niveaus)",
    ]
    cols2_nl = [0, 2, 4, 6, 7, 8]

    # Blad 3: Competentiebeschrijvingen
    ws3 = wb_src["3 Competence Statements"]
    rows3: List[Tuple] = []
    for r in range(2, ws3.max_row + 1):
        area_no = norm_num(ws3.cell(r, 1).value)
        comp_no = norm_num(ws3.cell(r, 4).value)
        sid     = str(ws3.cell(r, 7).value or "").strip()
        prof_s  = str(ws3.cell(r, 9).value or "").strip()
        if not sid:
            continue
        rows3.append((
            ws3.cell(r, 1).value,
            tr(core, f"digcomp.area.{area_no}.label"),
            ws3.cell(r, 2).value,
            tr(core, f"digcomp.area.{area_no}.description"),
            ws3.cell(r, 3).value,
            ws3.cell(r, 4).value,
            tr(core, f"digcomp.competence.{comp_no}.label"),
            ws3.cell(r, 5).value,
            tr(core, f"digcomp.competence.{comp_no}.description"),
            ws3.cell(r, 6).value,
            sid,
            tr(stmts, f"digcomp.statement.{sid}"),
            ws3.cell(r, 8).value,
            level_name_map.get(prof_s.lower(), "") if prof_s else "",
            prof_s,
            tr_ai(str(ws3.cell(r, 10).value or "")),
        ))

    hdrs3_bi = [
        "Nummer competentiegebied",
        "Naam competentiegebied", "Competence area name",
        "Omschrijving competentiegebied", "Competence area descriptor",
        "Nummer competentie",
        "Naam competentie", "Competence name",
        "Omschrijving competentie", "Competence descriptor",
        "ID competentiebeschrijving",
        "Competentiebeschrijving", "Competence statement",
        "Naam beheersingsniveau", "Proficiency level name",
        "AI-label",
    ]
    hdrs3_nl = [
        "Nummer competentiegebied",
        "Naam competentiegebied", "Omschrijving competentiegebied",
        "Nummer competentie", "Naam competentie", "Omschrijving competentie",
        "ID competentiebeschrijving", "Competentiebeschrijving",
        "Naam beheersingsniveau", "AI-label",
    ]
    cols3_nl = [0, 1, 3, 5, 6, 8, 10, 11, 13, 15]

    # Blad 4: Leerresultaten
    ws4 = wb_src["4 Learning Outcomes"]
    rows4: List[Tuple] = []
    for r in range(2, ws4.max_row + 1):
        area_no = norm_num(ws4.cell(r, 1).value)
        comp_no = norm_num(ws4.cell(r, 3).value)
        oid     = str(ws4.cell(r, 5).value or "").strip()
        prof_s  = str(ws4.cell(r, 7).value or "").strip()
        if not oid:
            continue
        rows4.append((
            ws4.cell(r, 1).value,
            tr(core, f"digcomp.area.{area_no}.label"),
            ws4.cell(r, 2).value,
            ws4.cell(r, 3).value,
            tr(core, f"digcomp.competence.{comp_no}.label"),
            ws4.cell(r, 4).value,
            oid,
            tr(outs, f"digcomp.outcome.{oid}"),
            ws4.cell(r, 6).value,
            level_name_map.get(prof_s.lower(), "") if prof_s else "",
            prof_s,
            tr_ksa(str(ws4.cell(r, 8).value or "")),
            tr_ai(str(ws4.cell(r, 9).value or "")),
        ))

    hdrs4_bi = [
        "Nummer competentiegebied",
        "Naam competentiegebied", "Competence area name",
        "Nummer competentie",
        "Naam competentie", "Competence name",
        "ID leerresultaat",
        "Leerresultaat", "Learning Outcome",
        "Beheersingsniveau", "Proficiency level",
        "Kennis, vaardigheid of attitude",
        "AI-label",
    ]
    hdrs4_nl = [
        "Nummer competentiegebied",
        "Naam competentiegebied",
        "Nummer competentie", "Naam competentie",
        "ID leerresultaat", "Leerresultaat",
        "Beheersingsniveau",
        "Kennis, vaardigheid of attitude",
        "AI-label",
    ]
    cols4_nl = [0, 1, 3, 4, 6, 7, 9, 11, 12]

    # Blad 5: Verklarende woordenlijst
    ws5 = wb_src["5 Glossary"]
    rows5: List[Tuple] = []
    for r in range(2, ws5.max_row + 1):
        term = ws5.cell(r, 1).value
        if not term:
            continue
        term_s = str(term).strip()
        slug = slugify_term(term_s)
        rows5.append((
            tr(gloss, f"digcomp.glossary.{slug}.label"),    # NL term
            term_s,                                          # EN term
            tr(gloss, f"digcomp.glossary.{slug}.definition"),  # NL toelichting
            str(ws5.cell(r, 2).value or ""),                  # EN explanation
        ))

    hdrs5_bi = ["Term (NL)", "Term (EN)", "Toelichting", "Explanation"]
    hdrs5_nl = ["Term", "Toelichting"]
    cols5_nl = [0, 2]

    # --- Helper om één variant te bouwen ---

    def _make_variant(bilingual: bool, out_xlsx: Path) -> None:
        wb = Workbook()
        wb.remove(wb.active)  # verwijder lege standaard-sheet

        # Leesmij als eerste blad
        if readme_nl_xlsx:
            copy_leesmij_sheet(wb, readme_nl_xlsx)

        def add_sheet(title: str, headers, rows_full, cols_nl):
            ws = wb.create_sheet(title)
            if bilingual:
                hdrs = headers[0]
                data = [list(row) for row in rows_full]
            else:
                hdrs = headers[1]
                data = [[row[i] for i in cols_nl] for row in rows_full]
            write_rows(ws, hdrs, data)

        add_sheet("1 Compgeb. & Competenties",
                  (hdrs1_bi, hdrs1_nl), rows1, cols1_nl)
        add_sheet("2 Beheersingsniveaus",
                  (hdrs2_bi, hdrs2_nl), rows2, cols2_nl)
        add_sheet("3 Competentiebeschrijvingen",
                  (hdrs3_bi, hdrs3_nl), rows3, cols3_nl)
        add_sheet("4 Leerresultaten",
                  (hdrs4_bi, hdrs4_nl), rows4, cols4_nl)
        add_sheet("5 Verklarende woordenlijst",
                  (hdrs5_bi, hdrs5_nl), rows5, cols5_nl)

        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_xlsx)
        print(f"[OK] Geschreven: {out_xlsx}")

    _make_variant(True,  out_dir / f"{base_name}_tweetalig.xlsx")
    _make_variant(False, out_dir / f"{base_name}_nl.xlsx")


# ----------------------------
# JSON-LD bouwen
# ----------------------------

def load_jsonld(path: Path) -> Dict:
    return json.loads(path.read_text(encoding="utf-8"))


# Mapping: huidige EN-achtige veldnaam → Nederlandse veldnaam
_FIELD_RENAME: Dict[str, str] = {
    "name_nl":                        "naam",
    "description_nl":                 "omschrijving",
    "four_levels_name_nl":            "vier_niveaus_naam",
    "four_levels_description_nl":     "vier_niveaus_omschrijving",
    "eight_levels_description_nl":    "acht_niveaus_omschrijving",
    "applies_to_nl":                  "doel",
    "four_levels_proficiency_name_nl":"vier_niveaus_beheersingsnaam",
}

ATTRIBUTION = {
    "Engelstalige_bron": (
        "Cosgrove, J. and Cachia, R., DigComp 3.0: European Digital Competence Framework "
        "– Fifth Edition, Publications Office of the European Union, Luxembourg, 2025, "
        "https://data.europa.eu/doi/10.2760/0001149, JRC144121."
    ),
    "Nederlandstalige_bron": (
        "Gorissen, P. en van Zanten, M., DigComp 3.0 Nederlands, "
        "iXperium Centre of Expertise Leren met ict, Nijmegen, Nederland, 2026. "
        "https://ixperium.nl"
    ),
    "licentie": (
        "Creative Commons Attribution 4.0 International (CC BY 4.0) "
        "https://creativecommons.org/licenses/by/4.0/"
    ),
}


def build_jsonld(
    src_jsonld: Path,
    out_jsonld: Path,
    tx: Dict[str, Dict[str, Dict[str, str]]],
) -> None:
    data = load_jsonld(src_jsonld)

    core   = tx.get("core-framework", {})
    levels = tx.get("levels", {})
    stmts  = tx.get("statements", {})
    outs   = tx.get("outcomes", {})
    gloss  = tx.get("glossary", {})

    level_name_map = build_levels_name_map(levels)

    def prof_uri_to_nl(uri: str) -> str:
        if not uri:
            return ""
        s = uri.split("/", 1)[-1].split("_", 1)[0]
        return level_name_map.get(s.lower(), "")

    graph = data.get("@graph", [])
    for node in graph:
        t   = node.get("@type")
        _id = node.get("@id", "")

        if t == "CompetenceArea":
            num = _id.split("/", 1)[-1]
            node["naam"]         = tr(core, f"digcomp.area.{num}.label") or node.get("name", "")
            node["omschrijving"] = tr(core, f"digcomp.area.{num}.description") or node.get("description", "")

        elif t == "Competence":
            cid = _id.split("/", 1)[-1]
            node["naam"]         = tr(core, f"digcomp.competence.{cid}.label") or node.get("name", "")
            node["omschrijving"] = tr(core, f"digcomp.competence.{cid}.description") or node.get("description", "")

        elif t == "CompetenceStatement":
            sid = _id.split("/", 1)[-1]
            node["omschrijving"] = tr(stmts, f"digcomp.statement.{sid}") or node.get("description", "")
            p = node.get("four_levels_proficiency_name")
            if isinstance(p, str) and p:
                node["vier_niveaus_beheersingsnaam"] = prof_uri_to_nl(p)
            if "ai_label" in node:
                node["ai_label"] = tr_ai(node["ai_label"])

        elif t == "LearningOutcome":
            oid = _id.split("/", 1)[-1]
            node["omschrijving"] = tr(outs, f"digcomp.outcome.{oid}") or node.get("description", "")
            p = node.get("four_levels_proficiency_name")
            if isinstance(p, str) and p:
                node["vier_niveaus_beheersingsnaam"] = prof_uri_to_nl(p)
            if "ai_label" in node:
                node["ai_label"] = tr_ai(node["ai_label"])
            if "type" in node:
                node["type"] = tr_ksa(node["type"])

        elif t == "ProficiencyLevel":
            eight = node.get("eight_levels_mapping")
            if isinstance(eight, float) and eight.is_integer():
                eight_s = str(int(eight))
            else:
                eight_s = str(eight).strip()
            node["vier_niveaus_naam"]         = tr(levels, f"digcomp.level.{eight_s}.four_level_name") or node.get("four_levels_name", "")
            node["vier_niveaus_omschrijving"] = tr(levels, f"digcomp.level.{eight_s}.four_level_description") or node.get("four_levels_description", "")
            node["acht_niveaus_omschrijving"] = tr(levels, f"digcomp.level.{eight_s}.eight_level_description") or node.get("eight_levels_description", "")
            node["doel"]                      = tr(levels, f"digcomp.level.{eight_s}.applies_to") or node.get("applies_to", "")

        elif t == "Glossary":
            # @id = "Glossary/<Term>"
            term_raw = _id.split("/", 1)[-1] if "/" in _id else _id
            slug = slugify_term(term_raw)
            nl_term = tr(gloss, f"digcomp.glossary.{slug}.label")
            nl_def  = tr(gloss, f"digcomp.glossary.{slug}.definition")
            if nl_term:
                node["term"]       = nl_term
            if nl_def:
                node["toelichting"] = nl_def

        # Verwijder eventuele oude _nl-suffixveldnamen die nog in de bron stonden
        for old_key in list(_FIELD_RENAME.keys()):
            if old_key in node:
                del node[old_key]

    # Bronvermelding toevoegen op het hoogste niveau (geldig in JSON-LD)
    data["bronvermelding"] = ATTRIBUTION

    out_jsonld.parent.mkdir(parents=True, exist_ok=True)
    out_jsonld.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] Geschreven: {out_jsonld}")


# ----------------------------
# CLI
# ----------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Bouw Nederlandstalige DigComp 3.0-leveringen vanuit vertaalde CSV's."
    )
    ap.add_argument("--repo-root",       default="..",
                    help="Map met digcomp3-l10n/ (standaard: ..)")
    ap.add_argument("--out-dir",         default="nl",
                    help="Uitvoermap (standaard: ./nl)")
    ap.add_argument("--build",           choices=["xlsx", "jsonld"], required=True,
                    help="Welk artefact te bouwen")
    ap.add_argument("--src-xlsx",        default="DigComp 3.0 Data Supplement 24 Nov 2025.xlsx",
                    help="Pad naar de Engelstalige bronspreadsheet")
    ap.add_argument("--src-jsonld",      default="DigComp 3.0 Data Supplement 24 Nov 2025.jsonld",
                    help="Pad naar het Engelstalige bron-JSON-LD-bestand")
    ap.add_argument("--readme-nl-xlsx",  default=None,
                    help="Pad naar ReadMe_NL.xlsx (bevat LEESMIJ-tabblad)")
    ap.add_argument("--base-name",       default="DigComp_3_0_Data_Supplement",
                    help="Basisnaam voor uitvoerbestanden (zonder extensie)")
    args = ap.parse_args()

    repo_root = Path(args.repo_root).resolve()
    out_dir   = Path(args.out_dir)
    if not out_dir.is_absolute():
        out_dir = (Path.cwd() / out_dir).resolve()

    tx = load_translations(repo_root)

    readme_nl = Path(args.readme_nl_xlsx) if args.readme_nl_xlsx else None

    if args.build == "xlsx":
        build_xlsx(
            src_xlsx      = Path(args.src_xlsx),
            out_dir       = out_dir,
            tx            = tx,
            readme_nl_xlsx= readme_nl,
            base_name     = args.base_name,
        )
    elif args.build == "jsonld":
        build_jsonld(
            src_jsonld = Path(args.src_jsonld),
            out_jsonld = out_dir / f"{args.base_name}_nl.jsonld",
            tx         = tx,
        )


if __name__ == "__main__":
    main()
